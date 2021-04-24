from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os, re
import time
import pandas as pd
import datetime
from datetime import date, datetime, timedelta
import openpyxl
# from webdriver_manager.chrome import ChromeDriverManager

t_date = date.today()
todays_date = t_date.strftime("%m-%d-%Y")

import logging
for handler in logging.root.handlers[:]:
    logging.root.removeHandler(handler)
os.makedirs('logs', exist_ok=True)
logging.basicConfig(filename=f'logs\\Slack_{todays_date}.log',
                    format='%(asctime)s %(message)s',
                    filemode='a',
                    level = logging.DEBUG)

t_date = date.today()
todays_date = t_date.strftime("%m-%d-%Y")

import configparser
config = configparser.ConfigParser()
config.read('Slack.config.ini')
channel_name = config['slck']['channel_name']

os.makedirs("SlackBot_Reports", exist_ok=True)
if not os.path.exists(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx'):
    excel = openpyxl.load_workbook('Slack_Bot_report.xlsx')
    excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')

# chromedriver = "C:\\Program Files\\Google\\chrome-driver\\chromedriver.exe"
chromedriver = config["chrome_path"]["loc"]
os.environ["webdriver.chrome.driver"] = chromedriver
browser = webdriver.Chrome(chromedriver)
browser.maximize_window()
browser.implicitly_wait(30)

def analyse_result(channel=None, id_container=None, cnt=0, msg_last=None):
    time.sleep(2)
    msg_len = browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]')
    id_container = [] if not id_container else id_container
    for i in range(len(msg_len)):
        id_val = browser.find_elements_by_xpath(f'//div[@class = "c-virtual_list__item"]')[i].get_attribute('id')
        if "divider" in id_val or "loadingSpinner" in id_val or "GAC84ESL9" in id_val:
            pass
        else:
            print(id_val)
            if id_val not in id_container:
                id_container.append(id_val)
                # time.sleep(2)
                msg = browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]')[i].text

                if channel.text == "churn_scapp":
                    churn_scrapp([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "sales_scapp":
                    sales_scapp([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "payments_scapp":
                    payments_scapp([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "reinstalls_scapp":
                    reinstalls_scapp([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "uninstalls_scapp":
                    uninstalls_scapp([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "uninstalls":
                    uninstalls([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "reinstalls":
                    reinstalls([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "uninstall_within_trial":
                    uninstall_within_trial([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "live_video_alert":
                    live_video_alert([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "sales":
                    sales([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "churn":
                    churns([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "payments":
                    payments([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "billing_page_actions":
                    billing_page([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "client_escalations":
                    client_escalation([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])
                elif channel.text == "remote_session_alerts":
                    remote_session_alerts([msg, time.strftime('%Y-%m-%d %H:%M', time.localtime(float(id_val)))])


            

    if time.strftime('%Y-%m-%d %H:%M', time.localtime(float(msg_last))) not in [ time.strftime('%Y-%m-%d %H:%M', time.localtime(float(i))) for i in id_container ]:
        # print(time.strftime('%Y-%m-%d %H:%M', time.localtime(float(msg_last))), [ time.strftime('%Y-%m-%d %H:%M', time.localtime(float(i))) for i in id_container ])
        body = browser.find_element_by_css_selector('body')
        body.send_keys(Keys.PAGE_DOWN)
        analyse_result(channel=channel, id_container=id_container, msg_last=msg_last, cnt=cnt+1)

def churns_sc(message):
    print(message)

def churn_scrapp(message):
    datas = [ '' for i in range(10)]
    if 'vajro notificationsapp' in message[0].lower():
        for msg in message[0].splitlines():
            if 'client' in msg.lower():
                if 'uninstalled' in msg.lower():
                    clnt = "uninstalled"
                    datas[0] = clnt
                elif 'paused' in msg.lower():
                    clnt = "uninstalled"
                    datas[0] = clnt

            if 'name' in msg.lower():
                name = msg.split(':')[-1]
                datas[1] = name
            if 'phone number' in msg.lower():
                phone = msg.split(':')[-1]
                datas[3] = phone
            if 'store url' in msg.lower():
                store_url = msg.split()[-1]
                datas[4] = store_url
            if 'app id' in msg.lower():
                appid = msg.split(':')[-1]
                datas[5] = appid
            if 'store:' in msg.lower():
                store = msg.split(':')[-1]
                datas[6] = store
            if 'plan' in msg.lower():
                plan = msg.split(':')[-1]
                datas[7] = plan
            if 'since payment' in msg.lower():
                payment = msg.split(':')[-1]
                datas[8] = payment  
            if 'username' in msg.lower():
                email = msg.split(':')[-1]
                datas[9] = email
    
        date_time = message[1]
        datas[2] = date_time
    
    if datas != [ '' for i in range(10) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['churn_scrapp']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')

def sales_scapp(message):
    datas = [ '' for i in range(8) ]
    print(message)
    if 'vajro notificationsapp' in message[0].lower():
        for index, msg in enumerate(message[0].splitlines()):
            if 'signed up' in msg.lower():
                if re.findall(r'https?://.*\.[a-z]{2,3}', msg.lower()):
                    store_url = re.findall(r'https?://.*\.[a-z]{2,3}', msg.lower())[0].split()[0]
                    datas[4] = store_url

                if 'shopify plan:' in msg.lower():
                    shopify_plan = msg.lower().split('shopify plan:')[-1]
                    datas[6] = shopify_plan

                if 'app id:' in msg.lower():
                    appid = re.findall(r'app id: [0-9]+', msg.lower())[0]
                    appid = appid.split(':')[-1]
                    datas[5] = appid

            if index == 2:
                client_name = msg
                datas[0] = client_name
            if index == 3:
                shopname = msg
                datas[7] = shopname

            if 'email' in msg.lower():
                try:
                    email = re.findall(r'[a-z0-9]+@[a-z\.]+.[a-z]{2,3}', msg.lower())[0].split()[0]
                    datas[3] = email
                except: pass

            if 'phone number:' in msg.lower():
                phone_number = msg.lower().split('phone number:')[-1]
                datas[2] = phone_number

        date_time = message[1]
        datas[1] = date_time

    # print(datas)
    if datas != [ '' for i in range(8) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['sales_scrapp']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')

def payments_scapp(message):
    datas = [ '' for i in range(9) ]
    if 'vajro notificationsapp' in message[0].lower():
        for index, msg in enumerate(message[0].splitlines()):
            
            if 'client has' in msg.lower():
                udr = re.findall(r'\[[a-z]+\]', msg.lower())
                udr = udr[0][1:-1] if udr else ''
                datas[0] = udr
                
                store_url = re.findall(r'https://.*\.[a-z]{2,3}', msg.lower())
                store_url = store_url[0] if store_url else ''
                datas[4] = store_url

                appid = re.findall(r'app id: [0-9]+', msg.lower())
                appid = appid[0].split(':')[-1] if appid else ''
                datas[5] = appid

                shopy_plan = re.findall(r'shopify plan: [a-z0-9\s]+', msg.lower())
                shopy_plan = shopy_plan[0].split(':')[-1] if shopy_plan else ''
                datas[8] = shopy_plan
            
            if index == 2:
                datas[1] = msg
            if index == 3:
                datas[6] = msg

            if 'subscribed plan' in msg.lower():
                plan = msg.split(':')[-1]
                datas[7] = plan

            if 'country' in msg.lower():
                country = msg.split(':')[-1]
                datas[3] = country

        date_time = message[1]
        datas[2] = date_time
    
    if datas != [ '' for i in range(9) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['payments_scrapp']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')

def reinstalls_scapp(message):
    datas = [ '' for i in range(4) ]
    if 'vajro notificationsapp' in message[0].lower():
        for index, msg in enumerate(message[0].splitlines()):
            if re.findall(r'https?://.*\.[a-z]{2,3}', msg.lower()):
                store_url = re.findall(r'https?://.*\.[a-z]{2,3}', msg.lower())[0]
                datas[1] = store_url
            try:
                appid = re.findall(r'[0-9]{4,6}', msg)
                datas[2] = appid[0]
            except: pass

            name = ' '.join(msg.split()[-3:-1])
            datas[3] = name
            
        date_time = message[1]
        datas[0] = date_time

    if datas != [ '' for i in range(4) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['reinstall_scapp']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')

def uninstalls_scapp(message):
    datas = [ '' for i in range(12) ]
    if 'vajro notificationsapp' in message[0].lower():
        for index, msg in enumerate(message[0].splitlines()):
            if 'client' in msg.lower():
                clnt = msg.split()[1]
                datas[0] = clnt
                clnt_url = msg.split()[-1]
                datas[1] = clnt_url

            if 'name' in msg.lower():
                name = msg.split(':')[-1]
                datas[2] = name

            if 'phone number' in msg.lower():
                phone = msg.split(':')[-1]
                datas[4] = phone

            if 'store url' in msg.lower():
                store_url = msg.split()[-1]
                datas[5] = store_url

            if 'app id' in msg.lower():
                appid = msg.split(':')[-1]
                datas[6] = appid

            if 'store:' in msg.lower():
                store = msg.split(':')[-1]
                datas[7] = store

            if 'vajro plan' in msg.lower():
                plan = msg.split(':')[-1]
                datas[8] = plan

            if 'no. of days' in msg.lower():
                no_of_days = msg.split(':')[-1]
                datas[9] = no_of_days

            if 'username' in msg.lower():
                email = msg.split(':')[-1]
                datas[10] = email

            if 'shopify plan' in msg.lower():
                shopify = msg.split(':')[-1]
                datas[11] = shopify
    
        date_time = message[1]
        datas[3] = date_time

    # print(datas)
    if datas != [ '' for i in range(12) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['uninstalls_scapp']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')

def uninstalls(message):
    datas = [ '' for i in range(10) ]
    if 'vajro notificationsapp' in message[0].lower():
        for msg in message[0].splitlines():

            if 'Name:' in msg:
                storename = msg.split(':')[-1]
                datas[1] = storename

            if 'app id' in msg.lower():
                appid = msg.split(':')[-1]
                datas[2] = appid

            if 'vajro plan' in msg.lower():
                vajro_plan = msg.split(':')[-1]
                datas[3] = vajro_plan

            if 'shopify plan' in msg.lower():
                shopify_plan = msg.split(':')[-1]
                datas[4] = shopify_plan

            if 'store url' in msg.lower():
                store_url = msg.split(':')[-1]
                datas[5] = 'https:' + store_url

            if 'no. of days' in msg.lower():
                no_of_days = msg.split(':')[-1]
                datas[6] = no_of_days

            if 'phone number' in msg.lower():
                phone_number = msg.split(':')[-1]
                datas[7] = phone_number
            
            if 'country' in msg.lower():
                country = msg.split(':')[-1]
                datas[8] = country

            if 'page link' in msg.lower():
                page_link = msg.split(':')[-1]
                datas[9] = 'https:' + page_link
        
        date_time = message[1]
        datas[0] = date_time
    
    if datas != [ '' for i in range(10) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['uninstalls']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')

def uninstall_within_trial(message):
    datas = [ '' for i in range(11) ]
    if 'vajro notificationsapp' in message[0].lower():
        for msg in message[0].splitlines():
            if 'client' in msg.lower():
                if 'uninstalled' in msg.lower():
                    client = "uninstalled"
                    datas[0] = client
                elif 'paused' in msg.lower():
                    client = "paused"
                    datas[0] = client
            
            if 'app id' in msg.lower():
                appid = msg.lower().split(':')[-1]
                datas[2] = appid

            if 'username' in msg.lower():
                email = msg.lower().split(':')[-1]
                datas[3] = email

            if 'store:' in msg.lower():
                storename = msg.lower().split(':')[-1]
                datas[4] = storename

            if 'store url' in msg.lower():
                storeurl = 'https:' + msg.lower().split(':')[-1]
                datas[5] = storeurl
            
            if 'phone number' in msg.lower():
                phone = msg.lower().split(':')[-1]
                datas[6] = phone

            if 'plan' in msg.lower():
                plan = msg.lower().split(':')[-1]
                datas[7] = plan

            if 'country' in msg.lower():
                country = msg.lower().split(':')[-1]
                datas[8] = country

            if 'since payment' in msg.lower():
                payment = msg.lower().split(':')[-1]
                datas[9] = payment

            if 'left in trial' in msg.lower():
                trial = msg.lower().split(':')[-1]
                datas[10] = trial

        date_time = message[1]
        datas[1] = date_time

    if datas != [ '' for i in range(11) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['uninstall_within_trail']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')

def reinstalls(message):
    datas = [ '' for i in range(4) ]
    if 'vajro notificationsapp' in message[0].lower():
        for msg in message[0].splitlines():
            if re.findall(r'https?://.*\.[a-z]{2,3}', msg.lower()):
                store_url = re.findall(r'https?://.*\.[a-z]{2,3}', msg.lower())[0]
                datas[1] = store_url

                try:
                    appid = re.findall(r'[0-9]{4,6}', msg)
                    datas[2] = appid[0]
                except: pass

                name = ' '.join(msg.split(appid[0])[-1].split()[:-1])
                datas[3] = name
        
        date_time = message[1]
        datas[0] = date_time

    if datas != [ '' for i in range(4) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['reinstalls']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')

def sales(txt):
    app_id = []
    store = []
    name = []
    email = []
    phone = []
    Country = []
    plan = []
    url = []
    datas = [ '' for i in range(9) ]
    # print('over')
    # time.sleep(3)
    # WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//span[@class = "p-channel_sidebar__name"][contains(text(), "sales")]'))).click()
    # time.sleep(4)
    # search_today()
    # length = len(browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]'))
    # print(length)
    # for i in range(length):
        # txt = browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]')[i].text
        # print(txt)
    a = txt[0].split('\n')
    try:
        apid = a[1].split(' ')
        for i in apid:
            if 'https' in i:
                url.append(i)
            else:
                url.append('No url')
        datas[2] = url[0]
    except:
        url.append('No url')
        datas[2] = url[0]
    try:
        plan.append(apid[-2]+' '+apid[-1])
        datas[8] = plan[0]
    except:
        plan.append('Plan')
        datas[8] = plan[0]
    try:
        app_id.append(apid[8].split('.')[0])
        datas[1] = app_id[0]
    except:
        app_id.append('No appid')
        datas[1] = app_id[0]
    try:
        name.append(a[2])
        datas[3] = name[0]
    except:
        name.append('No name')
        datas[3] = name[0]
    try:
        store.append(a[3])
        datas[7] = store[0]
    except:
        store.append('No Store')
        datas[7] = store[0]
    try:
        emal_num = a[4].split(' ')
        for i in emal_num:
            if "@" in i:
                email.append(i)
                datas[6] = email[0]
            else:
                email.append('No Email')
                datas[6] = email[0]
    except:
        email.append('No Email')
        datas[6] = email[0]
    # email.append(emal_num[0].split(':')[1])
    try:
        phone.append(emal_num[-1])
        datas[5] = phone[0]
    except:
        phone.append('No phone')
        datas[5] = phone[0]
    try:
        Country.append(a[5].split(":")[1])
        datas[4] = Country[0]
    except:
        Country.append('No Country')
        datas[4] = Country[0]

    datas[0] = txt[1]
    print(datas)
    # len of lists into var num
    # num= max(len(url), len(app_id), len(plan), len(name), len(store), len(email), len(phone), len(Country), len(dat))
    # # lists stored in fix_csv_len
    # for fix_csv_len in [url, app_id, plan, name, store, email, phone, Country, dat]:
    #     if len(fix_csv_len) < num: 
    #         for fix in range(num - len(fix_csv_len)):
    #             fix_csv_len.append(0)
    
    # date_time = txt[1]

    if datas != [ '' for i in range(9) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['sales']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
    
    # result = {"NAME":name, "DATE":dat, "COUNTRY": Country, "EMAIL": email, "PHONE NUMBER":phone, "URL":url, "APP ID":app_id, "STORE" : store, "PLAN": plan}
    # logging.info(f"{result}")
    # df = pd.DataFrame(result)
    # df.to_excel(f"SlackBot_Reports\\Sales_{todays_date}.xlsx", index=False)
    # writer = pd.ExcelWriter(f"Slack_Bot_Report_{todays_date}.xlsx", engine='xlsxwriter')
    # df.to_excel(writer, sheet_name='Sales_Report', index=False)
    # writer.save()

def search_today():
    i = 0
    while True:
        print(i)
        date_lable_element = WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,"//div[.='Loading history...']/following-sibling::div[1]/div/div/button[@aria-label='Jump to date']")))
        hist_date= date_lable_element.text
        print(hist_date, "history date")
        time.sleep(2)
        body = browser.find_element_by_css_selector('body')
        if i == 10:
            break
        else:
            body.send_keys(Keys.PAGE_UP)
        i+=1

def live_video_alert(txt):
    live_video = []
    dta = []
    store_url = []
    appid = []
    store_name = []
    target_audience = []
    fb_link = []
    sale_count = []
    datas = [ '' for i in range(9) ]
    # WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//span[@class = "p-channel_sidebar__name"][contains(text(), "live_video_alert")]'))).click()
    # time.sleep(3)
    # search_today()

    # length = len(browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]'))
    # print(length)

    # for li in range(length):
        # txt = browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]')[li].text
        # id_val = browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]')[li].get_attribute("id")
    
        # if "divider" in id_val or "loadingSpinner" in id_val:
        #     pass
        # else:
        #     print('epoch', id_val, time.localtime(float(id_val)))
    
        # print(txt)
    a = txt[0].split('\n')
    # print(a)
    try:
        for lvs in a:
            if "live video started" in lvs.lower():
                live = lvs.split(":")[1]
                live_video.append(live)
                datas[1] = live_video[0]
            
    except:
        live_video.append("No Live video data")
        datas[1] = live_video[0]
        

    try:
        for lvs in a:
            if "vajro notifications" in lvs.lower():
                dt = lvs.split(" ")[-2] + " " +lvs.split(" ")[-1]
                de = dt.replace("Today", f"{todays_date}")
                dta.append(de)
                datas[2] = dta[0]
    except:
        dta.append(f"{todays_date}")
        datas[2] = dta[0]

    try:
        for lvs in a:
            if "store url" in lvs.lower():
                url = lvs.split(" ")[-1]
                store_url.append(url)
                datas[3] = store_url[0]
    except:
        store_url.append("No URL")
        datas[3] = store_url[0]

    try:
        for lvs in a:
            if "app id" in lvs.lower():
                app = lvs.split(" ")[-1] 
                appid.append(app)
                datas[4] = appid[0]
    except:
        appid.append("No APP ID")
        datas[4] = appid[0]

    try:
        for lvs in a:
            if "store name" in lvs.lower():
                stre = lvs.split(":")[-1] 
                store_name.append(stre)
                datas[5] = store_name[0]
    except:
        store_name.append("No store name")
        datas[5] = store_name[0]

    try:
        for lvs in a:
            if "target audience" in lvs.lower():
                target = lvs.split(":")[-1] 
                target_audience.append(target)
                datas[6] = target_audience[0]
    except:
        target_audience.append("No Target Audience")
        datas[6] = target_audience[0]

    try:
        for lvs in a:
            if "facebook link" in lvs.lower():
                fb = lvs.split(" ")[-1] 
                fb_link.append(fb)
                datas[7] = fb_link[0]
    except:
        fb_link.append("No Facebook link")
        datas[7] = fb_link[0]

    try:
        for lvs in a:
            if "live sale count" in lvs.lower():
                sale_cnt = lvs.split(":")[-1] 
                sale_count.append(sale_cnt)
                datas[8] = sale_count[0]
    except:
        sale_count.append("No Live Sale Count")
        datas[8] = sale_count[0]
    
    datas[0] = txt[1]

    if datas != [ '' for i in range(9) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['live_video_alert']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
            print("datas[col-1]",datas[col-1])
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        #print("hjk")
    # num= max(len(live_video), len(dta), len(store_url), len(appid), len(store_name), len(target_audience), len(fb_link), len(sale_count))
    # # lists store_named in fix_csv_len
    # for fix_csv_len in [live_video, dta, store_url, appid, store_name, target_audience, fb_link, sale_count]:
    #     if len(fix_csv_len) < num: 
    #         for fix in range(num - len(fix_csv_len)):
    #             fix_csv_len.append("No Data")

    # video_alert_result = {
    #     "Live Video Started":live_video,
    #     "Date":dta,
    #     "Store URL":store_url,
    #     "APP ID":appid,
    #     "Store Name":store_name,
    #     "Target Audience":target_audience,
    #     "Facebook Link":fb_link,
    #     "Live Sale Count":sale_count
    # }
    # print(video_alert_result)

    # df = pd.DataFrame(video_alert_result)
    # df.to_excel(f"SlackBot_Reports\\Live_Video_Alert_{todays_date}.xlsx", index=False)

def churns(txt):
    client_pause = []
    client_name = []
    phone = []
    store_url = []
    app_id = []
    store_name = []
    vajro_plan = []
    no_of_days_since_payment = []
    country = []
    datas = [ '' for i in range(10) ]
    # WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//span[@class = "p-channel_sidebar__name"][contains(text(), "churn")]'))).click()
    # time.sleep(3)
    # search_today()
    # length = len(browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]'))
    # print(length)

    # for li in range(length):
    #     txt = browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]')[li].text
    a = txt[0].split('\n')
    # print(a)
    try:
        for client_pase in a:
            if "client paused" in client_pase.lower():
                # pause = client_pase.split(" ")[-1]
                client_pause.append("Pause")
                datas[1] = client_pause[0]
            if "client uninstalled" in client_pase.lower():
                client_pause.append("Uninstalled")
                datas[1] = client_pause[0]
    except:
        client_pause.append("No Data")
        datas[1] = client_pause[0]
    try:
        for client_pase in a:
            if "name" in client_pase.lower():
                name = client_pase.split(":")[-1]
                client_name.append(name)
                datas[2] = client_name[0]
                break
    except:
        client_name.append("No Data")
        datas[2] = client_name[0]
    try:
        for client_pase in a:
            if "phone number" in client_pase.lower():
                phne = client_pase.split(":")[-1]
                phone.append(phne)
                datas[3] = phone[0]
                break
    except:
        phone.append("No Data")
        datas[3] = phone[0]
    try:
        for client_pase in a:
            if "store url" in client_pase.lower():
                phne = client_pase.split(" ")[-1]
                store_url.append(phne)
                datas[4] = store_url[0]
                break
    except:
        store_url.append("No Data")
        datas[4] = store_url[0]
    try:
        for client_pase in a:
            if "app id" in client_pase.lower():
                apdi = client_pase.split(":")[-1]
                app_id.append(apdi)
                datas[5] = app_id[0]
                break
    except:
        app_id.append("No Data")
        datas[5] = app_id[0]
    try:
        for client_pase in a:
            if "store" in client_pase.lower():
                stre = client_pase.split(":")[-1]
                store_name.append(stre)
                datas[6] = store_name[0]
                break
    except:
        store_name.append("No Data")
        datas[6] = store_name[0]
    try:
        for client_pase in a:
            if "plan" in client_pase.lower():
                plan = client_pase.split(":")[-1]
                vajro_plan.append(plan)
                datas[7] = vajro_plan[0]
                break
    except:
        vajro_plan.append("No Data")
        datas[7] = vajro_plan[0]
    try:
        for client_pase in a:
            if "since payment" in client_pase.lower():
                since_pay = client_pase.split(":")[-1]
                no_of_days_since_payment.append(since_pay)
                datas[8] = no_of_days_since_payment[0]
                break
    except:
        no_of_days_since_payment.append("No Data")
        datas[8] = no_of_days_since_payment[0]
    try:
        for client_pase in a:
            if "country" in client_pase.lower():
                contry = client_pase.split(":")[-1]
                country.append(contry)
                datas[9] = country[0]
                break
    except:
        country.append("No Data")
        datas[9] = country[0]

    datas[0] = txt[1]

    if datas != [ '' for i in range(10) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['churns']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
    # num= max(len(client_pause), len(client_name), len(phone), len(store_url), len(app_id), len(store_name), len(vajro_plan), len(no_of_days_since_payment), len(country))
    # for fix_csv_len in [client_pause, client_name, phone, store_url, app_id, store_name, vajro_plan, no_of_days_since_payment, country]:
    #     if len(fix_csv_len) < num: 
    #         for fix in range(num - len(fix_csv_len)):
    #             fix_csv_len.append("No Data")

    # churn = {
    #     "Client Pause/Uninstalled":client_pause,
    #     "Client Name":client_name,
    #     "Phone Number":phone,
    #     "Store URL":store_url,
    #     "APP ID":app_id,
    #     "Store Name":store_name,
    #     "Vajro Plan":vajro_plan,
    #     "No of Days Since Payment":no_of_days_since_payment,
    #     "Country":country
    # }
    # print(churn)

    # df = pd.DataFrame(churn)
    # df.to_excel(f"SlackBot_Reports\\Churn_{todays_date}.xlsx", index=False)

def payments(txt):
    upgrade = []
    client_name = []
    dat = []
    country = []
    store_url = []
    app_id = []
    store_name = []
    vajro_plan = []
    shopify_plan = []
    datas = [ '' for i in range(10) ]
    # WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//span[@class = "p-channel_sidebar__name"][contains(text(), "payments")]'))).click()
    # time.sleep(3)
    # search_today()
    # length = len(browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]'))
    # print(length)

    # for li in range(length):
    #     txt = browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]')[li].text
    a = txt[0].split('\n')
    # print(a)
    try:
        for client_pase in a:
            if "upgrade" in client_pase.lower():
                upgrade.append("Upgrade")
                datas[1] = upgrade[0]
            if "downgrade" in client_pase.lower():
                upgrade.append("Downgrade")
                datas[1] = upgrade[0]
    except:
        pass
    try:
        for client_pase in a:
            if "uid" in client_pase.lower() and "at" in client_pase.lower():
                dte = client_pase.split("|")[-1].replace("Today", f"{todays_date}")
                # print(dte)
                dat.append(dte)
                datas[3] = dat[0]
    except:
        dat.append("No Data")
        datas[3] = dat[0]
    try:
        for client_pase in a:
            if "country" in client_pase.lower():
                conty = client_pase.split(":")[-1]
                # print(conty)
                country.append(conty)
                datas[4] = country[0]
    except:
        country.append("No Data")
        datas[4] = country[0]
    try:
        for client_pase in a:
            if "app id" in client_pase.lower():
                conty = client_pase.split(" ")
                # print(conty)
                shopify_plan.append(conty[-2]+ ' '+conty[-1])
                datas[9] = shopify_plan[0]
                for i in conty:   
                    if "https" in i.lower():
                        # print(i)
                        store_url.append(i)
                        datas[5] = store_url[0]
                    if "id" in i.lower():
                        ap = (conty.index(i)+1)
                        apid = conty[ap]
                        app_id.append(apid)
                        datas[6] = app_id[0]
    except:
        pass
    try:
        for client_pase in a:
            if "subscribed plan" in client_pase.lower():
                conty = client_pase.split(":")[-1]
                vajro_plan.append(conty)
                datas[8] = vajro_plan[0]
    except:
        vajro_plan.append("No Data")
        datas[8] = vajro_plan[0]
    try:
        client_name.append(a[2])
        datas[2] = client_name[0]
        store_name.append(a[3])
        datas[7] = store_name[0]
    except:
        client_name.append("No Data")
        datas[2] = client_name[0]
        store_name.append("No Data")
        datas[7] = store_name[0]
    
    datas[0] = txt[1]

    if datas != [ '' for i in range(10) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['payments']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        # num= max(len(upgrade), len(client_name), len(dat), len(country), len(store_url), len(app_id), len(store_name), len(vajro_plan), len(shopify_plan))
        # for fix_csv_len in [upgrade, client_name, dat, country, store_url, app_id, store_name, vajro_plan, shopify_plan]:
        #     if len(fix_csv_len) < num: 
        #         for fix in range(num - len(fix_csv_len)):
        #             fix_csv_len.append("No Data")

        # payment = {
        #     "Upgrade/Downgrade":upgrade,
        #     "Client Name":client_name,
        #     "Date":dat,
        #     "Country":country,
        #     "Store URL":store_url,
        #     "APP ID":app_id,
        #     "Store Name":store_name,
        #     "Vajro Plan":vajro_plan,
        #     "Shopify Plan":shopify_plan
        # }
        # print(payment)

        # df = pd.DataFrame(payment)
        # df.to_excel(f"SlackBot_Reports\\payment_{todays_date}.xlsx", index=False)

def billing_page(txt):
    page_visited = []
    store_url = []
    app_id = []
    store_name = []
    time_visited = []
    datas = [ '' for i in range(6) ]
    # WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//span[@class = "p-channel_sidebar__name"][contains(text(), "billing_page_actions")]'))).click()
    # time.sleep(3)
    # search_today()
    # length = len(browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]'))
    # print(length)

    # for li in range(length):
    #     txt = browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]')[li].text
    a = txt[0].split('\n')
    # print(a)
    if "vajro" not in a[0]:
        page_visited.append(a[0]) 
        datas[1] = page_visited[0]
    try:
        for client_pase in a:
            if "domain" in client_pase.lower():
                str_url = client_pase.split(" ")[-1]
                store_url.append(str_url)
                datas[2] = store_url[0]
                break
    except:
        store_url.append("No Data")
        datas[2] = store_url[0]
    try:
        for client_pase in a:
            if "app id" in client_pase.lower():
                str_url = client_pase.split(":")[-1]
                app_id.append(str_url)
                datas[3] = app_id[0]
                break
    except:
        app_id.append("No Data")
        datas[3] = app_id[0]
    try:
        for client_pase in a:
            if "name" in client_pase.lower():
                str_url = client_pase.split(":")[-1]
                store_name.append(str_url)
                datas[4] = store_name[0]
                break
    except:
        store_name.append("No Data")
        datas[4] = store_name[0]

    datas[0] = txt[1]

    if datas != [ '' for i in range(10) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['billing_page']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        # num= max(len(page_visited), len(store_url), len(app_id), len(store_name))
        # for fix_csv_len in [page_visited, store_url, app_id, store_name]:
        #     if len(fix_csv_len) < num: 
        #         for fix in range(num - len(fix_csv_len)):
        #             fix_csv_len.append("No Data")

        # billing_page_visited = {
        #     "Billing page visited/payment initiated":page_visited,
        #     "Store URL":store_url,
        #     "APP ID":app_id,
        #     "Store Name":store_name
        # }
        # print(billing_page_visited)

        # df = pd.DataFrame(billing_page_visited)
        # df.to_excel(f"SlackBot_Reports\\billing_page_visited_{todays_date}.xlsx", index=False)

def client_escalation(txt):
    store_name = []
    app_id = []
    vajro_plan = []
    shopify_plan = []
    store_url = []
    tenure_days_vajro = []
    mail = []
    intercom_link = []
    query = []
    datas = [ '' for i in range(10) ]
    # WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//span[@class = "p-channel_sidebar__name"][contains(text(), "client_escalations")]'))).click()
    # time.sleep(3)
    # search_today()
    # length = len(browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]'))
    # print(length)
    # for li in range(length):
        # txt = browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]')[li].text
    a = txt[0].split('\n')
    # print(a)
    try:
        for client_pase in a:
            if "storename" in client_pase.lower():
                contry = client_pase.split(":")[-1]
                store_name.append(contry)
                datas[1] = store_name[0]
                break
    except:
        store_name.append("No Data")
        datas[1] = store_name[0]
    try:
        for client_pase in a:
            if "app id" in client_pase.lower():
                contry = client_pase.split(":")[-1]
                app_id.append(contry)
                datas[2] = app_id[0]
                break
    except:
        app_id.append("No Data")
        datas[2] = app_id[0]
    try:
        for client_pase in a:
            if "vajro plan" in client_pase.lower():
                contry = client_pase.split(":")[-1]
                vajro_plan.append(contry)
                datas[3] = vajro_plan[0]
                break
    except:
        vajro_plan.append("No Data")
        datas[3] = vajro_plan[0]
    try:
        for client_pase in a:
            if "shopify plan" in client_pase.lower():
                contry = client_pase.split(":")[-1]
                shopify_plan.append(contry)
                datas[4] = shopify_plan[0]
                break
    except:
        shopify_plan.append("No Data")
        datas[4] = shopify_plan[0]
    try:
        for client_pase in a:
            if "store url" in client_pase.lower():
                contry = client_pase.split(" ")[-1]
                store_url.append(contry)
                datas[5] = store_url[0]
                break
    except:
        store_url.append("No Data")
        datas[5] = store_url[0]
    try:
        for client_pase in a:
            if "tenure days with vajro" in client_pase.lower():
                contry = client_pase.split(":")[-1]
                tenure_days_vajro.append(contry)
                datas[6] = tenure_days_vajro[0]
                break
    except:
        tenure_days_vajro.append("No Data")
        datas[6] = tenure_days_vajro[0]
    try:
        for client_pase in a:
            if "mailid" in client_pase.lower():
                contry = client_pase.split(":")[-1]
                mail.append(contry)
                datas[7] = mail[0]
                break
    except:
        mail.append("No Data")
        datas[7] = mail[0]
    try:
        for client_pase in a:
            if "intercom profile link" in client_pase.lower():
                contry = client_pase.split(":")[-1]
                intercom_link.append(contry)
                datas[8] = intercom_link[0]
                break
    except:
        intercom_link.append("No Data")
        datas[8] = intercom_link[0]
    try:
        for client_pase in a:
            if "query description" in client_pase.lower():
                contry = client_pase.split(":")[-1]
                query.append(contry)
                datas[9] = query[0]
                break
    except:
        query.append("No Data")
        datas[9] = query[0]

    datas[0] = txt[1]

    if datas != [ '' for i in range(10) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['client_escalation']
        row = churn_scrapp.max_row
        for col in range(1, churn_scrapp.max_column+1):
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        # num= max(len(store_name), len(app_id), len(vajro_plan), len(shopify_plan), len(store_url), len(tenure_days_vajro), len(mail), len(intercom_link), len(query))
        # for fix_csv_len in [mail, tenure_days_vajro, intercom_link, shopify_plan, store_url, app_id, store_name, vajro_plan, query]:
        #     if len(fix_csv_len) < num: 
        #         for fix in range(num - len(fix_csv_len)):
        #             fix_csv_len.append("No Data")

        # client_escalations = {
        #     "Store Name":store_name,
        #     "APP ID":app_id,
        #     "Vajro Plan":vajro_plan,
        #     "Shopify Plan":shopify_plan,
        #     "Store URL":store_url,
        #     "Tenure Days With Vajro":tenure_days_vajro,
        #     "Mail ID":mail,
        #     "Intercom Link":intercom_link,
        #     "Query Description":query
        # }

        # print(client_escalations)

        # df = pd.DataFrame(client_escalations)
        # df.to_excel(f"SlackBot_Reports\\client_escalations_{todays_date}.xlsx", index=False)


def remote_session_alerts(txt):
    dta = []
    schedule = []
    store_url = []
    app_id = []
    store_name = []
    intercom_link = []
    datas = [ '' for i in range(4) ]
    a = txt[0].split('\n')
    # print(a)
    try:
        for client_pase in a:
            if "vajro notifications" in client_pase.lower():
                dt = client_pase.split(" ")[-2] + " " +client_pase.split(" ")[-1]
                de = dt.replace("Today", f"{todays_date}")
                dta.append(de)
                datas[1] = dta[0]
    except:
        dta.append(f"{todays_date}")
        datas[1] = dta[0]
    try:
        for client_pase in a:
            if 'schedule' in client_pase.lower():
                # pause = client_pase.split(" ")[-1]
                schedule.append("Schedule")
                datas[2] = schedule[0]
            if 'immediate' in client_pase.lower():
                schedule.append("IMMEDIATE")
                datas[2] = schedule[0]
    except:
        schedule.append("No Data")
        datas[2] = schedule[0]

    try:
        for client_pase in a:
            if 'domain' in client_pase.lower():
                url = client_pase.split(":")[-1]
                store_url.append(url)
                datas[3] = store_url[0]
    except:
        store_url.append("No URL")
        datas[3] = store_url[0]

    try:
        for client_pase in a:
            if 'app id' in client_pase.lower():
                app = client_pase.split(" ")[-1] 
                app_id.append(app)
                datas[4] = app_id[0]
    except:
        app_id.append("No APP ID")
        datas[4] = app_id[0]

    try:
        for client_pase in a:
            if 'name' in client_pase.lower():
                stre = client_pase.split(":")[-1] 
                store_name.append(stre)
                datas[5] = store_name[0]
    except:
        store_name.append("No store name")
        datas[5] = store_name[0]

    try:
        for client_pase in a:
            if 'https' in client_pase.lower():
                contry = client_pase.split(":")[-1]
                intercom_link.append(contry)
                datas[6] = intercom_link[0]

    except:
        intercom_link.append("No Data")
        datas[6] = intercom_link[0]

    datas[0] = txt[1]
    #print(datas)
    if datas != [ '' for i in range(4) ]:
        excel = openpyxl.load_workbook(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        churn_scrapp = excel['remote_session_alerts']
        #print("churn",churn_scrapp)
        row = churn_scrapp.max_row
        #print("row",row)
        for col in range(1, churn_scrapp.max_column+1):
            #print('row',row)
            #print('col',col)
            #print('datas',datas[col-1])
            print(row, col, datas[col-1])
            churn_scrapp.cell(row=row+1, column=col).value = datas[col-1]
            #print(datas[col-1])
        excel.save(f'SlackBot_Reports\\Slack_Bot_report_{todays_date}.xlsx')
        #print("he")



def login():
    # try:
        # logging.info("RAPBot has started for the day to scrap information from sales channel")
    browser.get(config["slck"]["url"])
    WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//input[@name = "domain"]'))).send_keys(config["slck"]["workspace"])
    WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//button[@type = "submit"]'))).click()
    WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//input[@name = "email"]'))).send_keys(config["slck"]["uid"])
    WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//input[@name = "password"]'))).send_keys(config["slck"]["pwd"])
    WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//button[@type = "submit"]'))).click()
    time.sleep(5)

    channel = WebDriverWait(browser, 60).until(EC.element_to_be_clickable((By.XPATH,'//span[@class = "p-channel_sidebar__name"][contains(text(), "{0}")]'.format(channel_name))))
    print(channel.text)
    browser.execute_script("arguments[0].scrollIntoView();", channel)
    browser.execute_script("arguments[0].click();", channel)
    last_date_id = browser.find_elements_by_xpath('//*[@class="c-message_kit__gutter"]')[-1].get_attribute('id')
    msg_last = browser.find_elements_by_xpath('//div[@class = "c-virtual_list__item"]')[-1].get_attribute('id')
    time.sleep(15)
    analyse_result(channel=channel, msg_last=msg_last)
login()
